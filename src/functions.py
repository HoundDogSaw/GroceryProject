# Imports
from openpyxl.styles import Border, Side, Alignment, Font
import PySimpleGUI as sg
import sys
from constants import recipeList, staplesList


def build_header(wsh, title_name) -> None:
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(size=12, bold=True, name='Arial')
    border = Border(left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium'))
    new_row = wsh.max_row
    if title_name != "Produce":
        # wsh.insert_rows(new_row, 3)
        new_row += 4
    wsh.cell(column=1, row=new_row, value=title_name)
    wsh.row_dimensions[new_row].height = 14
    wsh.cell(column=1, row=new_row).alignment = alignment
    wsh.cell(column=1, row=new_row).border = border
    wsh.cell(column=1, row=new_row).font = font
    wsh.merge_cells(start_row=str(new_row), end_row=str(new_row), start_column='1', end_column='9')

    return


def build_list(wsh, ingredients, item_list) -> None:
    next_entry = max((c.row for c in wsh['A'] if c.value is not None))
    col = 1
    for r in ingredients.items():
        if r[0] in item_list:
            wsh.cell(column=col, row=next_entry + 1, value=r[0]).font = Font(size=9)
            wsh.cell(column=col + 1, row=next_entry + 1, value=r[1]).font = Font(size=9)
            col += 3
            if col == 10:
                next_entry += 1
                col = 1
    return


def build_menu():
    # GUI Menu window
    layout_menu = [[sg.Text('SOUP', text_color='black')],
                   [sg.Checkbox('Black Bean Soup', key='black_bean_soup', size=(18, 1)),
                    sg.Checkbox('Potato Soup', key='potato_soup', size=(18, 1)),
                    sg.Checkbox('French Onion Soup', key='french_onion_soup', size=(18, 1))],
                   [sg.Checkbox('Zuppa Toscana Soup', key='zuppa_toscana_soup', size=(18, 1)),
                    sg.Checkbox('Stuffed Pepper Soup', key='stuffed_pepper_soup', size=(18, 1)),
                    sg.Checkbox('Chicken Wild Rice Soup', key='chicken_wild_rice_soup', size=(22, 1))],
                   [sg.Text('MEAT', text_color='black')],
                   [sg.Checkbox('Steak', key='steak', size=(10, 1)),
                    sg.Checkbox('Burgers', key='burgers', size=(10, 1)),
                    sg.Checkbox('Pork Chops', key='pork_chops', size=(10, 1)),
                    sg.Checkbox('Chicken', key='chicken', size=(10, 1)), sg.Checkbox('Fish', key='fish', size=(10, 1))],
                   [sg.Text('MEXICAN', text_color='black')],
                   [sg.Checkbox('Steak Fajitas', key='steak_fajitas', size=(18, 1)),
                    sg.Checkbox('Chicken Fajitas', key='chicken_fajitas', size=(18, 1)),
                    sg.Checkbox('Tacos', key='tacos', size=(18, 1))],
                   [sg.Checkbox('Burrito Bowls', key='burrito_bowls', size=(18, 1)),
                    sg.Checkbox('Enchiladas', key='enchiladas', size=(18, 1))],
                   [sg.Text('SANDWICHES', text_color='black')],
                   [sg.Checkbox('BLT', key='blt', size=(18, 1)), sg.Checkbox('BBQ', key='bbq', size=(18, 1)),
                    sg.Checkbox('Lettuce Wraps', key='lettuce_wraps', size=(18, 1))],
                   [sg.Text('DISHES', text_color='black')],
                   [sg.Checkbox('Bean Dish', key='bean_dish', size=(18, 1)),
                    sg.Checkbox('Pot Pie', key='pot_pie', size=(18, 1)),
                    sg.Checkbox('White Chicken Chili', key='white_chicken_chili', size=(18, 1))],
                   [sg.Checkbox('Pasta', key='pasta', size=(18, 1)),
                    sg.Checkbox('Steak Stir Fry', key='steak_stir_fry', size=(18, 1)),
                    sg.Checkbox('Beef Stew', key='beef_stew', size=(18, 1))],
                   [sg.Checkbox('Gumbo', key='gumbo', size=(18, 1)),
                    sg.Checkbox('Basil Chicken Stir Fry', key='basil_chicken_stir_fry', size=(18, 1)),
                    sg.Checkbox('Pizza', key='pizza', size=(18, 1))],
                   [sg.Checkbox('Blackened Chicken', key='blackened_chicken', size=(18, 1))],
                   [sg.Submit(tooltip='Submit'), sg.Cancel()]]

    # GUI Staples window
    layout_staples = [[sg.Text('FRUIT', text_color='black')],
                      [sg.Checkbox('Strawberries', key='strawberries', size=(18, 1)),
                       sg.Checkbox('Raspberries', key='raspberries', size=(18, 1)),
                       sg.Checkbox('Grapes', key='grapes', size=(18, 1))],
                      [sg.Checkbox('Lemons', key='lemons', size=(18, 1)),
                       sg.Checkbox('Bananas', key='bananas', size=(18, 1)),
                       sg.Checkbox('Green Apples', key='green_apples', size=(18, 1))],
                      [sg.Checkbox('Limes', key='limes', size=(18, 1))],
                      [sg.Text('PRODUCE', text_color='black')],
                      [sg.Checkbox('Mini Cukes', key='mini_cukes', size=(18, 1)),
                       sg.Checkbox('Cherubs', key='cherubs', size=(18, 1))],
                      [sg.Text('MEAT', text_color='black')],
                      [sg.Checkbox('Hot Dogs', key='hot_dogs', size=(18, 1)),
                       sg.Checkbox('Bacon', key='bacon', size=(18, 1)),
                       sg.Checkbox('Sausage', key='sausage', size=(18, 1))],
                      [sg.Checkbox('Landjaegers', key='landjaegers', size=(18, 1)),
                       sg.Checkbox('Summer Sausage', key='summer_sausage', size=(18, 1))],
                      [sg.Text('CHEESE', text_color='black')],
                      [sg.Checkbox('Jalapeno Meunster', key='jalapeno_meunster', size=(18, 1)),
                       sg.Checkbox('Feta', key='feta', size=(18, 1)),
                       sg.Checkbox('Cheddar', key='cheddar', size=(18, 1))],
                      [sg.Checkbox('Swiss', key='swiss', size=(18, 1)),
                       sg.Checkbox('Mozzarella', key='mozzarella', size=(18, 1)),
                       sg.Checkbox('Colby', key='colby', size=(18, 1))],
                      [sg.Checkbox('Baby Swiss', key='baby_swiss', size=(18, 1)),
                       sg.Checkbox('Spreadable Cheese', key='spreadable_cheese', size=(18, 1)),
                       sg.Checkbox('Parmesean(block)', key='parmesean_block', size=(18, 1))],
                      [sg.Checkbox('Parmesean(grated)', key='parmesean_grated', size=(18, 1))],
                      [sg.Text('DAIRY', text_color='black')],
                      [sg.Checkbox('Milk', key='milk', size=(18, 1)), sg.Checkbox('Eggs', key='eggs', size=(18, 1)),
                       sg.Checkbox('Butter', key='butter', size=(18, 1))],
                      [sg.Checkbox('Yogurt(tub)', key='yogurt_tub', size=(18, 1)),
                       sg.Checkbox('Yogurt(pouch)', key='yogurt_pouch', size=(18, 1)),
                       sg.Checkbox('Yogurt(drink)', key='yogurt_drink', size=(18, 1))],
                      [sg.Checkbox('Sour Cream', key='sour_cream', size=(18, 1)),
                       sg.Checkbox('Creamer', key='creamer', size=(18, 1)),
                       sg.Checkbox('Yogurt(greek)', key='yogurt_greek', size=(18, 1))],
                      [sg.Checkbox('Cottage Cheese', key='cottage_cheese', size=(18, 1)),
                       sg.Checkbox('Orange Juice', key='orange_juice', size=(18, 1))],
                      [sg.Text('FROZEN', text_color='black')],
                      [sg.Checkbox('Waffles', key='waffles', size=(18, 1))],
                      [sg.Text('DRY GOODS', text_color='black')],
                      [sg.Checkbox('Coffee', key='coffee', size=(18, 1))],
                      [sg.Submit(tooltip='Submit'), sg.Cancel()]]

    # Menu window parameters
    window = sg.Window('Select menu items:', layout_menu, default_element_size=(40, 1), size=(550, 550))
    ing_needed = {}
    working = True
    while working:
        event, values = window.read()
        if event == 'Submit':
            for value in values:
                if values[value] and value in recipeList:
                    for ing, v in recipeList[value].items():
                        if ing not in ing_needed:
                            ing_needed[ing] = v
                        elif ing in ing_needed:
                            ing_needed[ing] += v
            window.close()
            working = False
        elif event == 'Cancel':
            window.close()
            sys.exit()
        else:
            window.close()
            sys.exit()

    window = sg.Window('Select additional grocery items:', layout_staples, default_element_size=(40, 1),
                       size=(550, 700))
    working = True
    while working:
        event, values = window.read()
        if event == 'Submit':
            for value in values:
                if values[value] and value in staplesList:
                    for ing, v in staplesList[value].items():
                        if ing not in ing_needed:
                            ing_needed[ing] = v
                        elif ing in ing_needed:
                            ing_needed[ing] += v
            window.close()
            working = False
        elif event == 'Cancel':
            window.close()
            sys.exit()
        else:
            window.close()
            sys.exit()

    return ing_needed

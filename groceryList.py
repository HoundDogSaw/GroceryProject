#!/usr/bin/env python
import PySimpleGUI as sg
import csv, os, sys
import pandas as pd
import openpyxl as op
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#Recipes
blt = {'bacon': 1, 'leaf lettuce': 1, 'tomato': 2}
steak = {'steak': 1, 'side vegetable': 1}
burgers = {'ground beef': 1, 'side vegetable': 1, 'cheddar': 1}
bean_dish = {'cannelini beans': 2, 'kidney beans': 1, 'baby spinach': 1, 'smoked sausage': 1, 'diced tomatoes': 1, 'garlic': 1, 'chicken boullion': 1}
enchiladas = {'corn tortillas': 1, 'ground beef': 1, 'black beans': 1, 'enchilada sauce': 1, 'diced chiles': 1, 'monterrey jack': 1, 'cilantro': 1, 'jalapeno': 1}
black_bean_soup = {'black beans': 1, 'salsa': 1, 'garlic': 1, 'chicken boullion': 1}
bbq = {'ground beef': 1, 'chicken gumbo soup': 1}
potato_soup = {'potatoes': 6, 'celery': 1, 'onion': 1, 'bacon': 1, 'ham': 1}
pot_pie = {'chicken breast': 1, 'carrots': 1, 'peas': 1, 'potatoes': 3, 'onion': 1, 'chicken boullion': 1}
steak_stir_fry = {'garlic': 1, 'sirloin': 1, 'onion': 1, 'cilantro': 1, 'soy sauce': 1, 'broccoli': 1, 'snow peas': 1}
blackened_chicken = {'chicken breast': 1, 'chicken boullion': 1, 'limes': 2, 'cilantro': 1, 'avocado': 1, 'yogurt(greek)': 1}
french_onion_soup = {'onion': 3, 'beef boullion': 1, 'worchestershire sauce': 1, 'swiss': 1}
white_chicken_chili = {'chicken breast': 1, 'onion': 1, 'garlic': 1, 'chicken boullion': 1, 'great northern beans': 2, 'diced chiles': 2, 'corn': 1, 'cilantro': 1, 'cream cheese': 1, 'whipping cream': 1}
zuppa_toscana_soup = {'bacon': 1, 'italian sausage': 1, 'garlic': 1, 'onion': 1, 'chicken boullion': 1, 'potatoes': 3, 'baby spinach': 1, 'whipping cream': 1}
stuffed_pepper_soup = {'ground beef': 1, 'onion': 1, 'garlic': 1, 'beef boullion': 1, 'tomato sauce(29oz)': 2, 'green peppers': 4, 'pre-cooked quinoa': 1, 'cheddar': 1}
fish = {'fish(pick)': 1, 'side vegetable': 1}
chicken = {'chicken(pick)': 1, 'side vegetable': 1}
pasta = {'noodles': 1, 'ground beef': 1, 'pasta sauce': 1, 'mozzarella': 1}
beef_stew = {'stew meat': 2, 'beef boullion': 1, 'onion': 1, 'carrot': 1, 'peas': 1, 'potatoes': 3}
tacos = {'ground beef': 1, 'tomatoes': 2, 'cilantro': 1, 'cheddar': 1, 'jalapeno': 1, 'taco shells': 1}
burrito_bowls = {'flank steak': 1, 'garlic': 1, 'limes': 4, 'soy sauce': 1, 'tomatoes': 2, 'cilantro': 1, 'black beans': 1, 'cheddar': 1, 'jalapeno': 1, 'avocado': 4}
butter_chicken = {'chicken breast': 1, 'lemons': 1, 'yogurt(greek)': 1, 'garlic': 1, 'jalapeno': 1, 'tomato sauce(8oz)': 1, 'whipping cream': 1, 'cilantro': 1}
pork_chops = {'pork chops': 1, 'side vegetable': 1}
steak_fajitas = {'sirloin': 1, 'red peppers': 1, 'green peppers': 1, 'yellow peppers': 1, 'onion': 1, 'fajita mix': 1, 'corn tortillas': 1}
chicken_fajitas = {'chicken breast': 1, 'red peppers': 1, 'green peppers': 1, 'yellow peppers': 1, 'onion': 1, 'fajita mix': 1, 'corn tortillas': 1}
chicken_wild_rice_soup = {'chicken breast': 1, 'carrots': 1, 'onion': 1, 'celery': 1, 'chicken boullion': 1, 'wild rice': 1, 'whipping cream': 1}
gumbo = {'smoked sausage': 2, 'onion': 1, 'green peppers': 2, 'celery': 1, 'garlic': 1, 'tomato paste': 1, 'diced tomatoes': 1, 'green onions': 1, 'chicken boullion': 1, 'okra': 1, 'shrimp': 1}
basil_chicken_stir_fry = {'chicken thigh': 1, 'green peppers': 2, 'jalapeno': 1, 'shallots': 1, 'basil': 1, 'fish sauce': 1}
lettuce_wraps = {'ground chicken': 1, 'garlic': 1, 'onion': 1, 'hoisin sauce': 1, 'ginger': 1, 'water chestnuts': 1, 'green onions': 1, 'head lettuce': 1}
pizza = {'pizza crust': 1, 'pizza sauce': 1, 'mozzarella': 1, 'pepperoni': 1, 'shitake': 2}

#List of Recipes
recipeList = {'blt': blt, 'steak': steak, 'burgers': burgers, 'bean_dish': bean_dish, 'enchiladas': enchiladas, 'black_bean_soup': black_bean_soup, 'potato_soup': potato_soup, 'french_onion_soup': french_onion_soup, 'zuppa_toscana_soup': zuppa_toscana_soup, 'stuffed_pepper_soup': stuffed_pepper_soup, 'chicken_wild_rice_soup': chicken_wild_rice_soup, 'pork_chops': pork_chops, 'chicken': chicken, 'fish': fish, 'steak_fajitas': steak_fajitas, 'chicken_fajitas': chicken_fajitas, 'tacos': tacos, 'burrito_bowls': burrito_bowls, 'bbq': bbq, 'lettuce_wraps': lettuce_wraps, 'pot_pie': pot_pie, 'white_chicken_chili': white_chicken_chili, 'pasta': pasta, 'steak_stir_fry': steak_stir_fry, 'beef_stew': beef_stew, 'gumbo': gumbo, 'basil_chicken_stir_fry': basil_chicken_stir_fry, 'pizza': pizza, 'blackened_chicken': blackened_chicken}


#Staples
mini_cukes = {'mini cukes': 1}
landjaegers = {'landjaegers': 1}
summer_sausage = {'summer sausage': 1}
jalapeno_meunster = {'jalapeno meunster': 1}
feta = {'feta': 1}
waffles = {'waffles': 1}
coffee = {'coffee': 1}
cheddar = {'cheddar': 1}
swiss = {'swiss': 1}
mozzarella = {'mozzarella': 1}
colby = {'colby': 1}
baby_swiss = {'baby swiss': 1}
spreadable_cheese = {'spreadable cheese': 1}
parmesean_block = {'parmesean(block)': 1}
parmesean_grated = {'parmesean(grated)': 1}
orange_juice = {'orange juice': 1}
milk = {'milk': 1}
eggs = {'eggs': 1}
butter = {'butter': 1}
cherubs = {'cherubs': 1}
green_apples = {'green apples': 5}
cottage_cheese = {'cottage cheese': 1}
yogurt_tub = {'yogurt(tub)': 1}
yogurt_pouch = {'yogurt(pouch)': 1}
yogurt_greek = {'yogurt(greek)': 1}
yogurt_drink = {'yogurt(drink)': 1}
strawberries = {'strawberries': 1}
raspberries = {'raspberries': 1}
grapes = {'grapes': 1}
lemons = {'lemons': 1}
limes = {'limes': 1}
bananas = {'bananas': 1}
sour_cream = {'sour cream': 1}
creamer = {'creamer': 1}
hot_dogs = {'hot dogs': 1}
bacon = {'bacon': 1}
sausage = {'sausage': 1}

#List of Staples
staplesList = {'strawberries': strawberries, 'raspberries': raspberries, 'grapes': grapes, 'lemons': lemons, 'bananas': bananas, 'green_apples': green_apples, 'limes': limes, 'mini_cukes': mini_cukes, 'cherubs': cherubs, 'hot_dogs': hot_dogs, 'bacon': bacon, 'sausage': sausage, 'landjaegers': landjaegers, 'summer_sausage': summer_sausage, 'jalapeno_meunster': jalapeno_meunster, 'feta': feta, 'cheddar': cheddar, 'swiss': swiss, 'mozzarella': mozzarella, 'colby': colby, 'baby_swiss': baby_swiss, 'spreadable_cheese': spreadable_cheese, 'parmesean_block': parmesean_block, 'parmesean_grated': parmesean_grated, 'milk': milk, 'eggs': eggs, 'butter': butter, 'yogurt_tub': yogurt_tub, 'yogurt_pouch': yogurt_pouch, 'yogurt_drink': yogurt_drink, 'yogurt_greek': yogurt_greek, 'sour_cream': sour_cream, 'creamer': creamer, 'cottage_cheese': cottage_cheese, 'orange_juice': orange_juice, 'waffles': waffles, 'coffee': coffee}

#Departments
fruit = ['strawberries', 'raspberries', 'grapes', 'lemons', 'limes', 'green apples', 'bananas']
produce1 = ['avocado', 'mini cukes', 'baby spinach', 'ginger', 'celery', 'carrots', 'jalapeno', 'green peppers', 'yellow peppers', 'red peppers', 'basil']
produce2 = ['side vegetable', 'green onion', 'onion', 'potatoes', 'cherubs', 'tomatoes', 'leaf lettuce', 'head lettuce', 'broccoli', 'cauliflower']
produce3 = ['garlic', 'cilantro', 'snow peas', 'okra', 'shallots', 'shitake', 'mushrooms']
meat1 = ['ground chicken', 'chicken breast', 'chicken thigh', 'chicken(pick)', 'bacon', 'smoked sausage', 'hot dogs', 'sausage']
meat2 = ['steak', 'pork chops', 'sirloin', 'flank steak', 'ground beef', 'italian sausage', 'stew meat', 'ham']
meat3 = ['pepperoni', 'shrimp', 'fish(pick)', 'landjaegers', 'summer sausage']
dairy1 = ['cream cheese', 'butter']
dairy2 = ['cheddar', 'swiss', 'monterrey jack', 'mozzarella', 'jalapeno meunster', 'feta', 'colby', 'baby swiss', 'spreadable cheese', 'parmesean(block)', 'parmesean(grated)']
dairy3 = ['whipping cream', 'cottage cheese', 'sour cream', 'creamer', 'yogurt(tub)', 'yogurt(pouch)', 'yogurt(drink)', 'yogurt(greek)', 'milk', 'eggs', 'orange juice']
frozen = ['peas', 'waffles']
gluten_free = ['noodles', 'pizza crust']
coffee_spices = ['coffee']
mexican = ['salsa', 'diced chiles', 'fajita mix', 'taco shells']
asian = ['fish sauce', 'hoisin sauce', 'water chestnuts', 'soy sauce']
soup = ['chicken boullion', 'beef boullion', 'chicken gumbo soup']
noodles_sauce = ['pasta sauce', 'pizza sauce']
beans_aisle = ['cannelini beans', 'kidney beans', 'black beans', 'great northern beans']
rice_tomatoes = ['wild rice', 'corn', 'tomato sauce(29oz)', 'pre-cooked quinoa', 'tomato paste', 'diced tomatoes']
condiments = ['worchestershire sauce']

#GUI Menu window
layout_menu = [[sg.Text('SOUP', text_color='black')],
          [sg.Checkbox('Black Bean Soup', key='black_bean_soup', size=(18, 1)), sg.Checkbox('Potato Soup', key='potato_soup', size=(18, 1)), sg.Checkbox('French Onion Soup', key='french_onion_soup', size=(18, 1))],
          [sg.Checkbox('Zuppa Toscana Soup', key='zuppa_toscana_soup', size=(18, 1)), sg.Checkbox('Stuffed Pepper Soup', key='stuffed_pepper_soup', size=(18, 1)), sg.Checkbox('Chicken Wild Rice Soup', key='chicken_wild_rice_soup', size=(22, 1))],
          [sg.Text('MEAT', text_color='black')],
          [sg.Checkbox('Steak', key='steak', size=(10, 1)), sg.Checkbox('Burgers', key='burgers', size=(10, 1)), sg.Checkbox('Pork Chops', key='pork_chops', size=(10, 1)), sg.Checkbox('Chicken', key='chicken', size=(10, 1)), sg.Checkbox('Fish', key='fish', size=(10, 1))],
          [sg.Text('MEXICAN', text_color='black')],
          [sg.Checkbox('Steak Fajitas', key='steak_fajitas', size=(18, 1)), sg.Checkbox('Chicken Fajitas', key='chicken_fajitas', size=(18, 1)), sg.Checkbox('Tacos', key='tacos', size=(18, 1))],
          [sg.Checkbox('Burrito Bowls', key='burrito_bowls', size=(18, 1)), sg.Checkbox('Enchiladas', key='enchiladas', size=(18, 1))],
          [sg.Text('SANDWICHES', text_color='black')],
          [sg.Checkbox('BLT', key='blt', size=(18, 1)), sg.Checkbox('BBQ', key='bbq', size=(18, 1)), sg.Checkbox('Lettuce Wraps', key='lettuce_wraps', size=(18, 1))],
          [sg.Text('DISHES', text_color='black')],
          [sg.Checkbox('Bean Dish', key='bean_dish', size=(18, 1)), sg.Checkbox('Pot Pie', key='pot_pie', size=(18, 1)), sg.Checkbox('White Chicken Chili', key='white_chicken_chili', size=(18, 1))],
          [sg.Checkbox('Pasta', key='pasta', size=(18, 1)), sg.Checkbox('Steak Stir Fry', key='steak_stir_fry', size=(18, 1)), sg.Checkbox('Beef Stew', key='beef_stew', size=(18, 1))],
          [sg.Checkbox('Gumbo', key='gumbo', size=(18, 1)), sg.Checkbox('Basil Chicken Stir Fry', key='basil_chicken_stir_fry', size=(18, 1)), sg.Checkbox('Pizza', key='pizza', size=(18, 1))],
          [sg.Checkbox('Blackened Chicken', key='blackened_chicken', size=(18, 1))],
          [sg.Submit(tooltip='Submit'), sg.Cancel()]]

#GUI Staples window
layout_staples = [[sg.Text('FRUIT', text_color='black')],
            [sg.Checkbox('Strawberries', key='strawberries', size=(18, 1)), sg.Checkbox('Raspberries', key='raspberries', size=(18, 1)), sg.Checkbox('Grapes', key='grapes', size=(18, 1))],
            [sg.Checkbox('Lemons', key='lemons', size=(18, 1)), sg.Checkbox('Bananas', key='bananas', size=(18, 1)), sg.Checkbox('Green Apples', key='green_apples', size=(18, 1))],
            [sg.Checkbox('Limes', key='limes', size=(18, 1))],
            [sg.Text('PRODUCE', text_color='black')],
            [sg.Checkbox('Mini Cukes', key='mini_cukes', size=(18, 1)), sg.Checkbox('Cherubs', key='cherubs', size=(18, 1))],
            [sg.Text('MEAT', text_color='black')],
            [sg.Checkbox('Hot Dogs', key='hot_dogs', size=(18, 1)), sg.Checkbox('Bacon', key='bacon', size=(18, 1)), sg.Checkbox('Sausage', key='sausage', size=(18, 1))],
            [sg.Checkbox('Landjaegers', key='landjaegers', size=(18, 1)), sg.Checkbox('Summer Sausage', key='summer_sausage', size=(18, 1))],
            [sg.Text('CHEESE', text_color='black')],
            [sg.Checkbox('Jalapeno Meunster', key='jalapeno_meunster', size=(18, 1)), sg.Checkbox('Feta', key='feta', size=(18, 1)), sg.Checkbox('Cheddar', key='cheddar', size=(18, 1))],
            [sg.Checkbox('Swiss', key='swiss', size=(18, 1)), sg.Checkbox('Mozzarella', key='mozzarella', size=(18, 1)), sg.Checkbox('Colby', key='colby', size=(18, 1))],
            [sg.Checkbox('Baby Swiss', key='baby_swiss', size=(18, 1)), sg.Checkbox('Spreadable Cheese', key='spreadable_cheese', size=(18, 1)), sg.Checkbox('Parmesean(block)', key='parmesean_block', size=(18, 1))],
            [sg.Checkbox('Parmesean(grated)', key='parmesean_grated', size=(18, 1))],
            [sg.Text('DAIRY', text_color='black')],
            [sg.Checkbox('Milk', key='milk', size=(18, 1)), sg.Checkbox('Eggs', key='eggs', size=(18, 1)), sg.Checkbox('Butter', key='butter', size=(18, 1))],
            [sg.Checkbox('Yogurt(tub)', key='yogurt_tub', size=(18, 1)), sg.Checkbox('Yogurt(pouch)', key='yogurt_pouch', size=(18, 1)), sg.Checkbox('Yogurt(drink)', key='yogurt_drink', size=(18, 1))],
            [sg.Checkbox('Sour Cream', key='sour_cream', size=(18, 1)), sg.Checkbox('Creamer', key='creamer', size=(18, 1)), sg.Checkbox('Yogurt(greek)', key='yogurt_greek', size=(18, 1))],
            [sg.Checkbox('Cottage Cheese', key='cottage_cheese', size=(18, 1)), sg.Checkbox('Orange Juice', key='orange_juice', size=(18, 1))],
            [sg.Text('FROZEN', text_color='black')],
            [sg.Checkbox('Waffles', key='waffles', size=(18, 1))],
            [sg.Text('DRY GOODS', text_color='black')],
            [sg.Checkbox('Coffee', key='coffee', size=(18, 1))],
            [sg.Submit(tooltip='Submit'), sg.Cancel()]]

#Menu window parameters
window = sg.Window('Select menu items:', layout_menu, default_element_size=(40, 1), size=(550, 450))

while True:
    event, values = window.read()
    if event == 'Submit':
        wb = load_workbook('test.xlsx')
        ws = wb.active
        menuRow = 1
        for value in values:
            if values[value] == True:
                print(value)
                ws.cell(column=5, row=menuRow, value=(value))
                menuRow = menuRow + 1
                wb.save('testGroceryList.xlsx')
                wb = load_workbook('testGroceryList.xlsx')
                ws = wb.active

                if value in recipeList:
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        recipe = locals()[value] #This converts value from a string to a variable
                        for ing, qty in recipe.items():
                            writer.writerow([ing, qty])

                '''if value == 'enchiladas':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in enchiladas.items():
                            writer.writerow([key, value])
                elif value == 'blt':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in blt.items():
                            writer.writerow([key, value])
                elif value == 'fish':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in fish.items():
                            writer.writerow([key, value])
                elif value == 'pizza':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in pizza.items():
                            writer.writerow([key, value])
                elif value == 'chicken_wild_rice_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in chicken_wild_rice_soup.items():
                            writer.writerow([key, value])
                elif value == 'lettuce_wraps':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in lettuce_wraps.items():
                            writer.writerow([key, value])
                elif value == 'basil_chicken_stir_fry':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in basil_chicken_stir_fry.items():
                            writer.writerow([key, value])
                elif value == 'gumbo':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in gumbo.items():
                            writer.writerow([key, value])
                elif value == 'blackened_chicken':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in blackened_chicken.items():
                            writer.writerow([key, value])
                elif value == 'steak':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in steak.items():
                            writer.writerow([key, value])
                elif value == 'burgers':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in burgers.items():
                            writer.writerow([key, value])
                elif value == 'black_bean_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in black_bean_soup.items():
                            writer.writerow([key, value])
                elif value == 'bean_dish':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in bean_dish.items():
                            writer.writerow([key, value])
                elif value == 'bbq':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in bbq.items():
                            writer.writerow([key, value])
                elif value == 'potato_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in potato_soup.items():
                            writer.writerow([key, value])
                elif value == 'pot_pie':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in pot_pie.items():
                            writer.writerow([key, value])
                elif value == 'steak_stir_fry':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in steak_stir_fry.items():
                            writer.writerow([key, value])
                elif value == 'french_onion_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in french_onion_soup.items():
                            writer.writerow([key, value])
                elif value == 'white_chicken_chili':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in white_chicken_chili.items():
                            writer.writerow([key, value])
                elif value == 'zuppa_toscana_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in zuppa_toscana_soup.items():
                            writer.writerow([key, value])
                elif value == 'stuffed_pepper_soup':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in stuffed_pepper_soup.items():
                            writer.writerow([key, value])
                elif value == 'chicken':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in chicken.items():
                            writer.writerow([key, value])
                elif value == 'pasta':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in pasta.items():
                            writer.writerow([key, value])
                elif value == 'beef_stew':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in beef_stew.items():
                            writer.writerow([key, value])
                elif value == 'tacos':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in tacos.items():
                            writer.writerow([key, value])
                elif value == 'burrito_bowls':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in burrito_bowls.items():
                            writer.writerow([key, value])
                elif value == 'pork_chops':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in pork_chops.items():
                            writer.writerow([key, value])
                elif value == 'steak_fajitas':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in steak_fajitas.items():
                            writer.writerow([key, value])
                elif value == 'chicken_fajitas':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in chicken_fajitas.items():
                            writer.writerow([key, value])'''

        window.close()
        break
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break

window = sg.Window('Select additional grocery items:', layout_staples, default_element_size=(40, 1), size=(550, 700))

while True:
    event, values = window.read()
    if event == 'Submit':
        for value in values:
            if values[value] == True:
                if value in staplesList:
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        staple = locals()[value] #This converts value from a string to a variable
                        for ing, qty in staple.items():
                            writer.writerow([ing, qty])
        window.close()
        '''break'''
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break

'''while True:
    event, values = window.read()
    if event == 'Submit':
        for value in values:
            if values[value] == True:
                if value == 'mini_cukes':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in mini_cukes.items():
                            writer.writerow([key, value])
                elif value == 'coffee':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in coffee.items():
                            writer.writerow([key, value])
                elif value == 'landjaegers':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in landjaegers.items():
                            writer.writerow([key, value])
                elif value == 'summer_sausage':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in summer_sausage.items():
                            writer.writerow([key, value])
                elif value == 'jalapeno_meunster':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in jalapeno_meunster.items():
                            writer.writerow([key, value])
                elif value == 'feta':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in feta.items():
                            writer.writerow([key, value])
                elif value == 'yogurt_greek':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in yogurt_greek.items():
                            writer.writerow([key, value])
                elif value == 'waffles':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in waffles.items():
                            writer.writerow([key, value])
                elif value == 'cheddar':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in cheddar.items():
                            writer.writerow([key, value])
                elif value == 'swiss':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in swiss.items():
                            writer.writerow([key, value])
                elif value == 'mozzarella':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in colby.items():
                            writer.writerow([key, value])
                elif value == 'colby':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in colby.items():
                            writer.writerow([key, value])
                elif value == 'baby_swiss':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in baby_swiss.items():
                            writer.writerow([key, value])
                elif value == 'spreadable_cheese':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in spreadable_cheese.items():
                            writer.writerow([key, value])
                elif value == 'parmesean_block':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in parmesean_block.items():
                            writer.writerow([key, value])
                elif value == 'parmesean_grated':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in parmesean_grated.items():
                            writer.writerow([key, value])
                elif value == 'milk':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in milk.items():
                            writer.writerow([key, value])
                elif value == 'eggs':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in eggs.items():
                            writer.writerow([key, value])
                elif value == 'butter':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in butter.items():
                            writer.writerow([key, value])
                elif value == 'cherubs':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in cherubs.items():
                            writer.writerow([key, value])
                elif value == 'green_apples':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in green_apples.items():
                            writer.writerow([key, value])
                elif value == 'yogurt_tub':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in yogurt_tub.items():
                            writer.writerow([key, value])
                elif value == 'yogurt_pouch':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in yogurt_pouch.items():
                            writer.writerow([key, value])
                elif value == 'yogurt_drink':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in yogurt_drink.items():
                            writer.writerow([key, value])
                elif value == 'cottage_cheese':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in cottage_cheese.items():
                            writer.writerow([key, value])
                elif value == 'orange_juice':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in orange_juice.items():
                            writer.writerow([key, value])
                elif value == 'strawberries':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in strawberries.items():
                            writer.writerow([key, value])
                elif value == 'raspberries':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in raspberries.items():
                            writer.writerow([key, value])
                elif value == 'grapes':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in grapes.items():
                            writer.writerow([key, value])
                elif value == 'bananas':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in bananas.items():
                            writer.writerow([key, value])
                elif value == 'lemons':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in lemons.items():
                            writer.writerow([key, value])
                elif value == 'limes':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in limes.items():
                            writer.writerow([key, value])
                elif value == 'hot_dogs':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in hot_dogs.items():
                            writer.writerow([key, value])
                elif value == 'bacon':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in bacon.items():
                            writer.writerow([key, value])
                elif value == 'sausage':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in sausage.items():
                            writer.writerow([key, value])
                elif value == 'sour_cream':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in sour_cream.items():
                            writer.writerow([key, value])
                elif value == 'creamer':
                    with open('output.csv', 'a+') as output:
                        writer = csv.writer(output)
                        for key, value in creamer.items():
                            writer.writerow([key, value])
        window.close()
        break
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break'''

def return_contents(file_name):
    with open(file_name) as infile:
        reader = csv.reader(infile)
        return list(reader)


groceryList = return_contents('output.csv')
#print(groceryList)

#This converts the count to an integer instead of string
for item in groceryList:
    item[1] = int(item[1])
#print(groceryList)

#This adds headers to the columns of the dataframe
title = ['ingredient', 'count']
df = pd.DataFrame(groceryList, columns = title)

#This groups the ingredients by header and sums the counts
df = df.groupby(by=['ingredient'])['count'].sum().reset_index()
df.to_csv(r'/home/pi/Programs/ingredients.csv', index=True)
#print(df)

wb = load_workbook('testGroceryList.xlsx')
ws = wb.active

'''wb = load_workbook('test.xlsx')
ws = wb.active'''

alignment = Alignment(horizontal='center', vertical='center')
font = Font(size=12, bold=True, name='Arial')
fontIngredient = Font(size=9)
whitefont = Font(size=10, name='Arial', color='FFFFFF')
border = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))
'''border2 = Border(top=Side(style='medium'))'''

#print(df)

for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fruit:
        next_entry = max((c.row for c in ws['A'] if c.value is not None))
        '''ws.row_dimensions[next_entry + 1].height = 11'''
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in produce1:
        next_entry = max((c.row for c in ws['A'] if c.value is not None))
        '''ws.row_dimensions[next_entry + 1].height = 11'''
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in produce2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        '''ws.row_dimensions[next_entry + 1].height = 11'''
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in produce3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        '''ws.row_dimensions[next_entry + 1].height = 11'''
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')

def build_department(section_title, name, added):
    newRowLocation = ws.max_row + 1
    ws.cell(column=1,row=newRowLocation, value=section_title)
    ws.insert_rows((ws.max_row), added)
    ws.row_dimensions[newRowLocation+added].height = 14
    ws.cell(column=1,row=newRowLocation+added).alignment = alignment
    ws.cell(column=1,row=newRowLocation+added).border = border
    '''ws.cell(column=3,row=newRowLocation+added).border = border2
    ws.cell(column=4,row=newRowLocation+added).border = border2
    ws.cell(column=5,row=newRowLocation+added).border = border2'''
    ws.cell(column=1,row=newRowLocation+added).font = font
    newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
    for r in dataframe_to_rows(df, index=False, header=False):
        if r[0] in name:
            '''ws.append(r).font = fontIngredient'''
            '''ws.row_dimensions[ws.max_row + 1, column=1].height = 10'''
            ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
            ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
    wb.save('testGroceryList.xlsx')

newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Meat')
ws.cell(column=4,row=newRowLocation, value='white letters')
ws.cell(column=5,row=newRowLocation, value='1')
ws.cell(column=7,row=newRowLocation, value='white letters')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 3)
ws.row_dimensions[newRowLocation+3].height = 14
ws.cell(column=1,row=newRowLocation+3).alignment = alignment
ws.cell(column=1,row=newRowLocation+3).border = border
'''ws.cell(column=3,row=newRowLocation+3).border = border2
ws.cell(column=4,row=newRowLocation+3).border = border2
ws.cell(column=5,row=newRowLocation+3).border = border2'''
ws.cell(column=1,row=newRowLocation+3).font = font
ws.cell(column=4,row=newRowLocation+3).font = whitefont
ws.cell(column=5,row=newRowLocation+3).font = whitefont
ws.cell(column=7,row=newRowLocation+3).font = whitefont
ws.cell(column=8,row=newRowLocation+3).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in meat1:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in meat2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in meat3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')

#build_department('Meat', meat, 2)

newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Dairy')
ws.cell(column=4,row=newRowLocation, value='white letters')
ws.cell(column=5,row=newRowLocation, value='1')
ws.cell(column=7,row=newRowLocation, value='white letters')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 3)
ws.row_dimensions[newRowLocation+3].height = 14
ws.cell(column=1,row=newRowLocation+3).alignment = alignment
ws.cell(column=1,row=newRowLocation+3).border = border
'''ws.cell(column=3,row=newRowLocation+3).border = border2
ws.cell(column=4,row=newRowLocation+3).border = border2
ws.cell(column=5,row=newRowLocation+3).border = border2'''
ws.cell(column=1,row=newRowLocation+3).font = font
ws.cell(column=4,row=newRowLocation+3).font = whitefont
ws.cell(column=5,row=newRowLocation+3).font = whitefont
ws.cell(column=7,row=newRowLocation+3).font = whitefont
ws.cell(column=8,row=newRowLocation+3).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in dairy1:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in dairy2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in dairy3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')

#build_department('Dairy', dairy, 2)
build_department('Frozen', frozen, 3)
build_department('Gluten Free', gluten_free, 2)
#build_department('Dry Goods', dry_goods, 1)

newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Dry Goods')
ws.insert_rows((ws.max_row), 1)
ws.row_dimensions[newRowLocation+1].height = 14
ws.cell(column=1,row=newRowLocation+1).alignment = alignment
ws.cell(column=1,row=newRowLocation+1).border = border
'''ws.cell(column=3,row=newRowLocation+1).border = border2
ws.cell(column=4,row=newRowLocation+1).border = border2
ws.cell(column=5,row=newRowLocation+1).border = border2'''
ws.cell(column=1,row=newRowLocation+1).font = font
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in coffee_spices:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in mexican:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in asian:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in soup:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in noodles_sauce:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in beans_aisle:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in rice_tomatoes:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')

newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Cereal/Snacks')
ws.insert_rows((ws.max_row), 3)
ws.row_dimensions[newRowLocation+3].height = 16
ws.cell(column=1,row=newRowLocation+3).alignment = alignment
ws.cell(column=1,row=newRowLocation+3).border = border
'''ws.cell(column=3,row=newRowLocation+3).border = border2
ws.cell(column=4,row=newRowLocation+3).border = border2
ws.cell(column=5,row=newRowLocation+3).border = border2'''
ws.cell(column=1,row=newRowLocation+3).font = font
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
wb.save('testGroceryList.xlsx')

newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Condiments')
ws.insert_rows((ws.max_row), 2)
ws.row_dimensions[newRowLocation+2].height = 16
ws.cell(column=1,row=newRowLocation+2).alignment = alignment
ws.cell(column=1,row=newRowLocation+2).border = border
'''ws.cell(column=3,row=newRowLocation+2).border = border2
ws.cell(column=4,row=newRowLocation+2).border = border2
ws.cell(column=5,row=newRowLocation+2).border = border2'''
ws.cell(column=1,row=newRowLocation+2).font = font
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in condiments:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')
wb.close()

#Converts .xlsx document to .pdf document
os.system('libreoffice --headless --convert-to pdf:calc_pdf_Export --outdir /home/pi/Programs/ /home/pi/Programs/testGroceryList.xlsx')

#This will print the list of ingredients in grocery list form
'''os.system('lpr -P HP_Officejet_Pro_8620_728EA3_ testGroceryList.pdf')'''

#This removes all files once the script completes
os.remove('output.csv')
os.remove('ingredients.csv')
os.remove('testGroceryList.xlsx')
'''os.remove('testGroceryList.pdf')'''




#!/usr/bin/env python
import PySimpleGUI as sg
import os, sys, platform
if platform.system() == 'Windows':
    import win32api
    import win32print
    from win32com import client
    import psutil
    import xlwings
import time
import pandas as pd
import openpyxl as op
import fooddata, layouts
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#Menu window parameters
window = sg.Window('Select menu items:', layouts.layout_menu, default_element_size=(40, 1), resizable=True)

#Menu window and menu sorting
ing_needed = {}
while True:
    event, values = window.read()
    if event == 'Submit':
        wb = load_workbook('listTemplate.xlsx')
        ws = wb.active
        menuRow = 1
        for value in values:
            if values[value] and value in fooddata.recipeList:
                #Adds menu to upper right of grocery list
                ws.cell(column=5, row=menuRow, value=(value))
                menuRow = menuRow + 1
                wb.save('testGroceryList.xlsx')
                wb = load_workbook('testGroceryList.xlsx')
                ws = wb.active
                #Organizes ingredients and sums like values
                for ing, v in fooddata.recipeList[value].items():
                    if ing not in ing_needed:
                        ing_needed[ing] = v
                    elif ing in ing_needed:
                        ing_needed[ing] += v
        window.close()
        break
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break

#Staples window parameters
window = sg.Window('Select additional grocery items:', layouts.layout_staples, default_element_size=(40, 1), resizable=True)

#Staples window and selection sorting
while True:
    event, values = window.read()
    if event == 'Submit':
        for value in values:
            if values[value] and value in fooddata.staplesList:
                for ing, v in fooddata.staplesList[value].items():
                    if ing not in ing_needed:
                        ing_needed[ing] = v
                    elif ing in ing_needed:
                        ing_needed[ing] += v
        window.close()
        break
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break

#Costco window parameters
window = sg.Window('Select items from CostCo:', layouts.layout_costco, default_element_size=(40, 1), resizable=True)

#CostCo window and selection sorting
while True:
    event, values = window.read()
    if event == 'Submit':
        for value in values:
            if values[value] and value in fooddata.costcoList:
                for ing, v in fooddata.costcoList[value].items():
                    if ing not in ing_needed:
                        ing_needed[ing] = v
                    elif ing in ing_needed:
                        ing_needed[ing] += v
        window.close()
        break
    elif event == 'Cancel':
        window.close()
        sys.exit()
        break
    else:
        window.close()
        sys.exit
        break
    
#This block converts ing_needed from a dictionary to dataframe and adds headers
title = ['ingredient', 'count']
df = pd.DataFrame(list(ing_needed.items()), columns=title)

#Loads pre-formatted Excel sheet to begin grocery list population
wb = load_workbook('testGroceryList.xlsx')
ws = wb.active

#Excel sheet formatting parameters
alignment = Alignment(horizontal='center', vertical='center')
font = Font(size=12, bold=True, name='Arial')
fontIngredient = Font(size=8)
whitefont = Font(size=10, name='Arial', color='FFFFFF')
border = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

#Fruit and produce organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.fruit:
        next_entry = max((c.row for c in ws['A'] if c.value is not None))
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.produce1:
        next_entry = max((c.row for c in ws['A'] if c.value is not None))
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.produce2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.produce3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient

#Function for building generic departments
def build_department(section_title, name, added):
    newRowLocation = ws.max_row + 1
    ws.cell(column=1,row=newRowLocation, value=section_title)
    ws.insert_rows((ws.max_row), added)
    ws.row_dimensions[newRowLocation+added].height = 14
    ws.cell(column=1,row=newRowLocation+added).alignment = alignment
    ws.cell(column=1,row=newRowLocation+added).border = border
    ws.cell(column=1,row=newRowLocation+added).font = font
    newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
    for r in dataframe_to_rows(df, index=False, header=False):
        if r[0] in fooddata.name:
            ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
            ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient

#More Excel formatting (Adds white letters in blank cells to facilitate .max_row function in 3 different columns)
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Meat')
ws.cell(column=4,row=newRowLocation, value='white letters')
ws.cell(column=5,row=newRowLocation, value='1')
ws.cell(column=7,row=newRowLocation, value='white letters')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 2)
ws.row_dimensions[newRowLocation+2].height = 14
ws.cell(column=1,row=newRowLocation+2).alignment = alignment
ws.cell(column=1,row=newRowLocation+2).border = border
ws.cell(column=1,row=newRowLocation+2).font = font
ws.cell(column=4,row=newRowLocation+2).font = whitefont
ws.cell(column=5,row=newRowLocation+2).font = whitefont
ws.cell(column=7,row=newRowLocation+2).font = whitefont
ws.cell(column=8,row=newRowLocation+2).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Meat organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.meat1:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.meat2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.meat3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient

#Even more Excel formatting (Adds white letters in blank cells to facilitate .max_row function in 3 different columns)
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Dairy')
ws.cell(column=4,row=newRowLocation, value='white letters')
ws.cell(column=5,row=newRowLocation, value='1')
ws.cell(column=7,row=newRowLocation, value='white letters')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 1)
ws.row_dimensions[newRowLocation+1].height = 14
ws.cell(column=1,row=newRowLocation+1).alignment = alignment
ws.cell(column=1,row=newRowLocation+1).border = border
ws.cell(column=1,row=newRowLocation+1).font = font
ws.cell(column=4,row=newRowLocation+1).font = whitefont
ws.cell(column=5,row=newRowLocation+1).font = whitefont
ws.cell(column=7,row=newRowLocation+1).font = whitefont
ws.cell(column=8,row=newRowLocation+1).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Dairy organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.dairy1:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.dairy2:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.dairy3:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        next_value = max((c.row for c in ws['H'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=8, row=next_value + 1, value=r[1]).font = fontIngredient

#Frozen and Toiletries formatting
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Frozen')
ws.cell(column=7,row=newRowLocation, value='Toiletries')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 1)
ws.row_dimensions[newRowLocation+1].height = 14
ws.cell(column=1,row=newRowLocation+1).alignment = alignment
ws.cell(column=1,row=newRowLocation+1).border = border
ws.cell(column=1,row=newRowLocation+1).font = font
ws.cell(column=7,row=newRowLocation+1).alignment = alignment
ws.cell(column=7,row=newRowLocation+1).border = border
ws.cell(column=7,row=newRowLocation+1).font = font
ws.cell(column=8,row=newRowLocation+1).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Frozen and Toiletries organization
for r in dataframe_to_rows(df, index=False, header=False):
        if r[0] in fooddata.frozen:
            ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
            ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.toiletries:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient

#Gluten Free formatting
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Gluten Free')
ws.insert_rows((ws.max_row), 1)
ws.row_dimensions[newRowLocation+1].height = 14
ws.cell(column=1,row=newRowLocation+1).alignment = alignment
ws.cell(column=1,row=newRowLocation+1).border = border
ws.cell(column=1,row=newRowLocation+1).font = font
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Gluten Free organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.gluten_free:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient

#Dry Goods and CostCo formatting
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Dry Goods')
ws.cell(column=7,row=newRowLocation, value='CostCo')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 1)
ws.row_dimensions[newRowLocation+1].height = 14
ws.cell(column=1,row=newRowLocation+1).alignment = alignment
ws.cell(column=1,row=newRowLocation+1).border = border
ws.cell(column=1,row=newRowLocation+1).font = font
ws.cell(column=7,row=newRowLocation+1).alignment = alignment
ws.cell(column=7,row=newRowLocation+1).border = border
ws.cell(column=7,row=newRowLocation+1).font = font
ws.cell(column=8,row=newRowLocation+1).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Dry Goods organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.coffee_spices:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.mexican:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.asian:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.soup:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.noodles_sauce:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.beans_aisle:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.rice_tomatoes:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
        
#CostCo organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.costco:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient

#Cereal/Snacks and Liquor formatting
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Cereal/Snacks')
ws.cell(column=4,row=newRowLocation, value='white letters')
ws.cell(column=5,row=newRowLocation, value='1')
ws.cell(column=7,row=newRowLocation, value='Liquor')
ws.cell(column=8,row=newRowLocation, value='1')
ws.insert_rows((ws.max_row), 2)
ws.row_dimensions[newRowLocation+2].height = 16
ws.cell(column=1,row=newRowLocation+2).alignment = alignment
ws.cell(column=1,row=newRowLocation+2).border = border
ws.cell(column=1,row=newRowLocation+2).font = font
ws.cell(column=4,row=newRowLocation+2).font = whitefont
ws.cell(column=5,row=newRowLocation+2).font = whitefont
ws.cell(column=7,row=newRowLocation+2).alignment = alignment
ws.cell(column=7,row=newRowLocation+2).border = border
ws.cell(column=7,row=newRowLocation+2).font = font
ws.cell(column=8,row=newRowLocation+2).font = whitefont
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')

#Cereal organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.cereal:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient

#Snacks organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.snacks:
        next_entry = max((c.row for c in ws['D'] if c.value is not None))
        next_value = max((c.row for c in ws['E'] if c.value is not None))
        ws.cell(column=4, row=next_entry + 1, value=r[0]).font = fontIngredient
        ws.cell(column=5, row=next_value + 1, value=r[1]).font = fontIngredient
        
#Liquor organization
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.liquor:
        next_entry = max((c.row for c in ws['G'] if c.value is not None))
        ws.cell(column=7, row=next_entry + 1, value=r[0]).font = fontIngredient

#Condiments formatting and organization
newRowLocation = ws.max_row + 1
ws.cell(column=1,row=newRowLocation, value='Condiments')
ws.insert_rows((ws.max_row), 2)
ws.row_dimensions[newRowLocation+2].height = 16
ws.cell(column=1,row=newRowLocation+2).alignment = alignment
ws.cell(column=1,row=newRowLocation+2).border = border
ws.cell(column=1,row=newRowLocation+2).font = font
newRowLocation = ws.merge_cells(start_row=str(ws.max_row), end_row=str(ws.max_row), start_column='1', end_column='2')
for r in dataframe_to_rows(df, index=False, header=False):
    if r[0] in fooddata.condiments:
        ws.cell(row=ws.max_row + 1, column=1, value=r[0]).font = fontIngredient
        ws.cell(row=ws.max_row, column=2, value=r[1]).font = fontIngredient
wb.save('testGroceryList.xlsx')
wb.close()

#Converts .xlsx document to .pdf document
'''os.system('libreoffice --headless --convert-to pdf:calc_pdf_Export --outdir /home/pi/Programs/Grocery_List/ /home/pi/Programs/Grocery_List/testGroceryList.xlsx')'''

#This will print the list of ingredients in grocery list form
if platform.system() == 'Linux':
    os.system('libreoffice --headless --convert-to pdf:calc_pdf_Export --outdir /home/pi/Programs/Grocery_List/ /home/pi/Programs/Grocery_List/testGroceryList.xlsx')
    os.system('lpr -P HP_Officejet_Pro_8620_728EA3_ testGroceryList.pdf')
elif platform.system() == 'Windows':
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False
    Workbook = app.Workbooks.Open(r'C:\Users\timhoff\Desktop\Personal\GroceryProject-Grocery-List-1.0\GroceryProject-Grocery-List-1.0\testGroceryList.xlsx')
    try:
        Workbook.ActiveSheet.ExportAsFixedFormat(0, r'C:\Users\timhoff\Desktop\Personal\GroceryProject-Grocery-List-1.0\GroceryProject-Grocery-List-1.0\testGroceryList.pdf')
    except Exception as e:
        print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
        print(str(e))
    finally:
        #win32api.ShellExecute (0,"print",r'C:\Users\timhoff\Desktop\Personal\GroceryProject-Grocery-List-1.0\GroceryProject-Grocery-List-1.0\testGroceryList.pdf', win32print.GetDefaultPrinter(),".",0)'''
        os.startfile(r'C:\Users\timhoff\Desktop\Personal\GroceryProject-Grocery-List-1.0\GroceryProject-Grocery-List-1.0\testGroceryList.pdf', "print")
        time.sleep(5)
        for p in psutil.process_iter(): #Close Acrobat after printing the PDF
            if 'AcroRd' in str(p):
                p.kill()
        Workbook.Close()
'''else platform.system() == 'Darwin':
    continue'''

#This removes temporary files once the script completes
'''os.remove('testGroceryList.xlsx')
os.remove('testGroceryList.pdf')'''